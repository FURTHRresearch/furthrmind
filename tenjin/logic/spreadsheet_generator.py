import datetime
import hashlib
import re
import time
from io import BytesIO
from itertools import zip_longest

import bson
import openpyxl
from bson.objectid import ObjectId
from mongoengine.queryset.base import Q
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from openpyxl.worksheet import worksheet

from tenjin.authentication import Authentication
from tenjin.database.db import get_db
from tenjin.execution.create import Create
from tenjin.execution.update import Update
from tenjin.file.file_storage import FileStorage
from tenjin.logic.change_detection import hasChanged
from tenjin.mongo_engine.Link import Link

from tenjin.web.researchitems import get_item_for_dashboard
from tenjin.web.helper.filter_utils import get_linked_items_recursive
import asyncio

class SpreadsheetGenerator:
    # noinspection PyDefaultArgument
    def __init__(self, exp_list=[], sample_list=[], research_item_list=[], dataview=None, user_id=None,
                 spreadsheet_id=None, template_id=None, force_template=False, save=True):
        if type(user_id) is str:
            user_id = ObjectId(user_id)
        self.user_id = user_id
        self.spreadsheet_id = spreadsheet_id
        self.template_id = template_id
        self.template = None
        if self.template_id:
            self.template: dict = self.getData("SpreadSheet", {"_id": self.template_id}, findOne=True,
                                               authCheck=True)
        self.force_template = force_template
        self.save = save
        self.exp_list: list = exp_list
        self.exp_id_list = [exp["_id"] for exp in exp_list]
        self.sample_list = sample_list
        self.sample_id_list = [sample["_id"] for sample in sample_list]
        self.research_item_list = research_item_list
        self.research_item_id_list = [research_item["_id"] for research_item in research_item_list]

        total_id_list = []
        total_id_list.extend(self.exp_id_list)
        total_id_list.extend(self.sample_id_list)
        total_id_list.extend(self.research_item_id_list)
        self.total_id_list = total_id_list

        links = Link.objects(Q(DataID1__in=total_id_list) | Q(DataID2__in=total_id_list))
        exp_id_list_to_extended = []
        sample_id_list_to_extended = []
        research_item_id_list_to_extended = []

        for link in links:
            if link.DataID1 not in total_id_list:
                if link.Collection1 == "Experiment":
                    exp_id_list_to_extended.append(link.DataID1)
                elif link.Collection1 == "Sample":
                    sample_id_list_to_extended.append(link.DataID1)
                elif link.Collection1 == "ResearchItem":
                    research_item_id_list_to_extended.append(link.DataID1)
            if link.DataID2 not in total_id_list:
                if link.Collection2 == "Experiment":
                    exp_id_list_to_extended.append(link.DataID2)
                elif link.Collection2 == "Sample":
                    sample_id_list_to_extended.append(link.DataID2)
                elif link.Collection2 == "ResearchItem":
                    research_item_id_list_to_extended.append(link.DataID2)
        if exp_id_list_to_extended:
            exp_list = self.getData("Experiment", {"_id": {"$in": exp_id_list_to_extended}})
            self.exp_list.extend(exp_list)
        if sample_id_list_to_extended:
            sample_list = self.getData("Sample", {"_id": {"$in": sample_id_list_to_extended}})
            self.sample_list.extend(sample_list)
        if research_item_id_list_to_extended:
            item_list = self.getData("ResearchItem", {"_id": {"$in": research_item_id_list_to_extended}})
            self.research_item_list.extend(item_list)

        self.dataview = dataview
        self.dataview_id = dataview["_id"] if dataview is not None else None
        self.std_fields_dataview = ["Name", "Short ID", "Type"]

        self.md5 = SpreadsheetGenerator.create_md5([],
                                                   exp_id_list=self.exp_id_list, sample_id_list=self.sample_id_list,
                                                   research_item_id_list=self.research_item_id_list,
                                                   dataview_id=self.dataview_id)

        self.fs = FileStorage(get_db())

        self.file_id = None
        self.outputBytes = None
        self.spreadsheet: dict = {}

        self.fieldDataList = []
        self.fieldDataIDList = []
        self.fieldList = []

        self.fieldDataMapping = {}
        self.fieldMapping = {}
        self.unitMapping = {}
        self.comboBoxEntryMapping = {}
        self.notebookMapping = {}
        self.linkMapping = {}
        self.linkTargetDataMapping = {}
        self.calculationMapping = {}

        self.rawDataList = []
        self.exp_id_rawdata_mapping = {}
        self.sample_id_rawdata_mapping = {}
        self.research_item_id_rawdata_mapping = {}
        self.rawDataIDList = []
        self.columnList = []
        self.columnIDList = []
        self.columnMapping = {}

        if self.dataview:
            find_dict = {"DataViewID": self.dataview_id}
            spreadsheet: dict = self.getData("SpreadSheet", find_dict, findOne=True)
            if spreadsheet:
                self.spreadsheet = spreadsheet
                self.file_id = self.spreadsheet["FileID"]
                self.update_workbook_for_dataview()
            else:
                self.create_workbook_for_dataview()

        else:
            start = time.time()
            # set expFieldDataList and expFieldDataIDList and fieldDataMapping
            self.getAllFieldDataList()
            # set expFieldList and fieldMapping
            self.getExpFieldList()

            # create comboBoxEntryMapping and completes fieldDataMapping and fieldMapping
            self.getComboBoxEntryFieldDataAndFieldRecursive(self.fieldDataIDList)

            self.getAllNotebook()
            self.getAllUnit()
            self.getAllLink()
            self.getAllCalculation()

            # set rawData and columnIDList
            self.getRawData()

            # columnList will be set only if new xlsx must be created.
            exist = False
            numberOfExpChanged = None
            if not self.force_template:
                exist, numberOfExpChanged = self.checkForExistingSpreadsheets()

            if exist:
                self.updateSheetNameYourWorkspace()
                changeString = self.checkForDataUpdates()
                if changeString != "Nothing":
                    self.updateXLSXAllExp(changeString)

            else:
                self.getRawDataColumnList()
                self.createXLSXAllItems()

    def getData(self, collection, findDict, findOne=False, authCheck=False):

        # Direct way
        if not authCheck:
            if findOne:
                result = get_db().db[collection].find_one(findDict)
            else:
                result = list(get_db().db[collection].find(findDict))
            return result
        else:
            result = list(get_db().db[collection].find(findDict))
            if collection == "User":
                for entry in result:
                    entry["Password"] = None

            if result:
                result = Authentication.filter_list_get(
                    collection, result, self.user_id, get_db())

            if findOne:
                return result[0] if result else None

            return result

    def checkForExistingSpreadsheets(self):
        numberOfExpChanged = False
        if self.spreadsheet_id:
            findDict = {"_id": self.spreadsheet_id}
            spreadsheet: dict = self.getData("SpreadSheet", findDict, findOne=True, authCheck=True)
            if not spreadsheet:
                return False, numberOfExpChanged
            self.spreadsheet = spreadsheet

            self.file_id = spreadsheet["FileID"]
            outputBytes = self.fs.get_file(self.file_id)
            self.outputBytes = outputBytes
            if len(spreadsheet["ExperimentIDList"]) != len(self.exp_id_list):
                self.deleteAllSheets()
                numberOfExpChanged = True
            return True, numberOfExpChanged
        else:
            findDict = {"MD5": self.md5}
            spreadsheet = self.getData("SpreadSheet", findDict, findOne=True)
            if spreadsheet:
                self.spreadsheet = spreadsheet
                self.spreadsheet_id = spreadsheet["_id"]
                self.file_id = spreadsheet["FileID"]
                outputBytes = self.fs.get_file(self.file_id)
                self.outputBytes = outputBytes
                return True, numberOfExpChanged
            return False, numberOfExpChanged

    @staticmethod
    def create_md5(id_list: list, exp_id_list=(), sample_id_list=(),
                   research_item_id_list=(), dataview_id=None):

        id_list.extend(exp_id_list)
        id_list.extend(sample_id_list)
        id_list.extend(research_item_id_list)
        if dataview_id:
            id_list.append(dataview_id)
        id_list = sorted(id_list)
        md5 = hashlib.md5(str(id_list).encode()).hexdigest()
        return md5

    def deleteAllSheets(self):
        wb = openpyxl.load_workbook(self.outputBytes)
        sheets = wb.sheetnames
        for sheetName in sheets:
            if sheetName == "Your workspace":
                continue
            try:
                sheet = wb[sheetName]
                wb.remove(sheet)
            except:
                pass

    def checkForDataUpdates(self):
        from tenjin.mongo_engine.RawData import RawData
        # webDataCalcIDList = list(self.webDataCalcMapping.keys())
        return_string = "Nothing"
        check = hasChanged(self.exp_id_list, self.spreadsheet["LastUpdate"],
                           get_db())
        if check:
            return_string = "Metadata"
        check = hasChanged(self.sample_id_list, self.spreadsheet["LastUpdate"],
                           get_db())
        if check:
            return_string = "Metadata"
        check = hasChanged(self.research_item_id_list,
                           self.spreadsheet["LastUpdate"],
                           get_db())
        if check:
            return_string = "Metadata"
        check = hasChanged(self.fieldDataIDList,
                           self.spreadsheet["LastUpdate"],
                           get_db())
        if check:
            return_string = "Metadata"
        if self.rawDataList:
            check = hasChanged([self.rawDataIDList], self.spreadsheet["LastUpdate"],
                               get_db())
            if check:
                if return_string == "Metadata":
                    return_string = "All"
                else:
                    return_string = "RawData"

        # find newly created raw_data...
        # get_db().db["RawData"].find({"$and": [{"Date": {"$gte": self.spreadsheet["LastUpdate"]}},
        #                                       {"$or": [
        #                                           {"ExpID": {"$in": self.exp_id_list}},
        #                                           {"SampleID": {"$in": self.sample_id_list}},
        #                                           {"ResearchItemID": {"$in": self.research_item_id_list}},
        #                                       ]}]})
        raw_data = RawData.objects((Q(ExpID__in=self.exp_id_list) | Q(SampleID__in=self.sample_id_list) |
                                    Q(ResearchItemID__in=self.research_item_id_list)) &
                                   Q(Date__gte=self.spreadsheet["LastUpdate"])).first()
        if raw_data:
            return_string = "RawData"

        if return_string != "Nothing":
            return return_string

        # noinspection DuplicatedCode
        if self.fieldMapping:
            fieldIDList = list(self.fieldMapping.keys())
            check = hasChanged(fieldIDList,
                               self.spreadsheet["LastUpdate"], get_db())
            if check:
                return "Metadata"
        if self.comboBoxEntryMapping:
            comboBoxEntryIDList = list(self.comboBoxEntryMapping.keys())
            check = hasChanged(comboBoxEntryIDList,
                               self.spreadsheet["LastUpdate"], get_db())
            if check:
                return "Metadata"
        # noinspection DuplicatedCode
        if self.unitMapping:
            unitIDList = list(self.unitMapping.keys())
            check = hasChanged(unitIDList, self.spreadsheet["LastUpdate"], get_db())
            if check:
                return "Metadata"
        if self.notebookMapping:
            notebookIDList = list(self.notebookMapping.keys())
            check = hasChanged(notebookIDList, self.spreadsheet["LastUpdate"],
                               get_db())
            if check:
                return "Metadata"
        if self.linkMapping:
            linkIDList = list(self.linkMapping.keys())
            check = hasChanged(linkIDList, self.spreadsheet["LastUpdate"],
                               get_db())
            if check:
                return "Metadata"
        return return_string

    def createXLSXAllItems(self):
        """
        returns bytes-like object, expName
        """
        if self.template:
            template_file_id = self.template["FileID"]
            template_bytes = self.fs.get_file(template_file_id)
            template_bytes_io = BytesIO(template_bytes)
            wb = openpyxl.load_workbook(template_bytes_io)
            self.remove_sheets(wb, "All")

        else:
            wb = openpyxl.Workbook()
            sheet = wb["Sheet"]
            sheet.title = "Your workspace"

        # noinspection DuplicatedCode
        expList = self.createListExp()
        self.createSheetExperimentalMetaDataSheet(wb, expList)

        sampleList = self.createListSample()
        self.createSheetSampleMetaDataSheet(wb, sampleList)

        researchList = self.createListResearchItems()
        self.createSheetResearchItemMetaDataSheet(wb, researchList)

        comboDataDict = self.createDictComboData()
        self.createSheetComboEntryMetaDataSheet(wb, comboDataDict)

        self.getRawDataColumnList()

        raw_data_sheet_names = []
        for exp in self.exp_list:
            raw_data_list = self.exp_id_rawdata_mapping.get(exp["_id"], [])
            for raw_data in raw_data_list:
                name = raw_data["Name"]
                if not name.lower().startswith("data table"):
                    name = f"Data table {name}"

                if name in raw_data_sheet_names:
                    name = f"{name} {exp['Name']}"
                    counter = 1
                    while name in raw_data_sheet_names:
                        name = f"{name}_{counter}"
                        counter += 1

                raw_data_sheet_names.append(name)

                if len(name) > 31:
                    name = name[0:31]
                replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                for char in replace_chars:
                    name = name.replace(char, "")
                self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                        name)
        for sample in self.sample_list:
            raw_data_list = self.sample_id_rawdata_mapping.get(sample["_id"], [])
            for raw_data in raw_data_list:
                name = raw_data["Name"]
                if not name.lower().startswith("data table"):
                    name = f"Data table {name}"

                if name in raw_data_sheet_names:
                    name = f"{name} {sample['Name']}"
                    counter = 1
                    while name in raw_data_sheet_names:
                        name = f"{name}_{counter}"
                        counter += 1

                raw_data_sheet_names.append(name)

                if len(name) > 31:
                    name = name[0:31]
                replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                for char in replace_chars:
                    name = name.replace(char, "")
                self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                        name)

        for researchitem in self.research_item_list:
            raw_data_list = self.research_item_id_rawdata_mapping.get(researchitem["_id"], [])
            for raw_data in raw_data_list:
                name = raw_data["Name"]
                if not name.lower().startswith("data table"):
                    name = f"Data table {name}"

                if name in raw_data_sheet_names:
                    name = f"{name} {researchitem['Name']}"
                    counter = 1
                    while name in raw_data_sheet_names:
                        name = f"{name}_{counter}"
                        counter += 1

                raw_data_sheet_names.append(name)

                if len(name) > 31:
                    name = name[0:31]
                replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                for char in replace_chars:
                    name = name.replace(char, "")
                self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                        name)

        outputIO = BytesIO()
        wb.save(outputIO)
        outputBytes = outputIO.getvalue()
        self.outputBytes = outputBytes

        fileName = ""
        for exp in self.exp_list:
            fileName += f"{exp['Name']}_"
        fileName += "spreadsheet.xlsx"
        self.file_id = self.fs.put(outputBytes, fileName=fileName)
        spreadsheetDict = {
            "ExperimentIDList": self.exp_id_list,
            "SampleIDList": self.sample_id_list,
            "ResearchItemIDList": self.research_item_id_list,
            "DataViewID": None,
            "LastUpdate": datetime.datetime.now(datetime.UTC),
            "FileID": self.file_id,
            "MD5": self.md5,
        }
        if self.save:
            _id = Create.create("SpreadSheet", spreadsheetDict, get_db(), self.user_id)
            self.spreadsheet_id = _id
            self.spreadsheet = get_db().db["SpreadSheet"].find_one({"_id": _id})

    def updateSheetNameYourWorkspace(self):
        oldBytesIO = BytesIO(self.outputBytes)
        wb = openpyxl.load_workbook(oldBytesIO, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if "Sheet" in sheets:
            wb = openpyxl.load_workbook(oldBytesIO)
            sheet = wb["Sheet"]
            sheet.title = "Your workspace"

            newBytesIO = BytesIO()
            wb.save(newBytesIO)
            wb.close()
            self.outputBytes = newBytesIO.getvalue()
            self.fs.put(self.outputBytes, self.file_id)
            Update.update("SpreadSheet", "LastUpdate", datetime.datetime.now(datetime.UTC),
                          self.spreadsheet["_id"],
                          get_db(), self.user_id)
        else:
            wb.close()

    def updateXLSXAllExp(self, changeString):

        oldBytesIO = BytesIO(self.outputBytes)
        wb = openpyxl.load_workbook(oldBytesIO)

        self.remove_sheets(wb, changeString)

        if changeString in ["Metadata", "All"]:
            expList = self.createListExp()
            self.createSheetExperimentalMetaDataSheet(wb, expList)

            sampleList = self.createListSample()
            self.createSheetSampleMetaDataSheet(wb, sampleList)

            riList = self.createListResearchItems()
            self.createSheetResearchItemMetaDataSheet(wb, riList)

            comboDataDict = self.createDictComboData()
            self.createSheetComboEntryMetaDataSheet(wb, comboDataDict)

        if changeString in ["RawData", "All"]:
            self.getRawDataColumnList()
            raw_data_sheet_names = []
            for exp in self.exp_list:
                raw_data_list = self.exp_id_rawdata_mapping.get(exp["_id"], [])
                for raw_data in raw_data_list:
                    name = raw_data["Name"]
                    if not name.lower().startswith("data table"):
                        name = f"Data table {name}"

                    if name in raw_data_sheet_names:
                        name = f"{name} {exp['Name']}"
                        counter = 1
                        while name in raw_data_sheet_names:
                            name = f"{name}_{counter}"
                            counter += 1

                    raw_data_sheet_names.append(name)

                    if len(name) > 31:
                        name = name[0:31]

                    replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                    for char in replace_chars:
                        name = name.replace(char, "")

                    self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                            name)
            for sample in self.sample_list:
                raw_data_list = self.sample_id_rawdata_mapping.get(sample["_id"], [])
                for raw_data in raw_data_list:
                    name = raw_data["Name"]
                    if not name.lower().startswith("data table"):
                        name = f"Data table {name}"

                    if name in raw_data_sheet_names:
                        name = f"{name} {sample['Name']}"
                        counter = 1
                        while name in raw_data_sheet_names:
                            name = f"{name}_{counter}"
                            counter += 1

                    raw_data_sheet_names.append(name)

                    if len(name) > 31:
                        name = name[0:31]
                    replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                    for char in replace_chars:
                        name = name.replace(char, "")
                    self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                            name)
            for researchitem in self.research_item_list:
                raw_data_list = self.research_item_id_rawdata_mapping.get(researchitem["_id"], [])
                for raw_data in raw_data_list:
                    name = raw_data["Name"]
                    if not name.lower().startswith("data table"):
                        name = f"Data table {name}"

                    if name in raw_data_sheet_names:
                        name = f"{name} {researchitem['Name']}"
                        counter = 1
                        while name in raw_data_sheet_names:
                            name = f"{name}_{counter}"
                            counter += 1

                    raw_data_sheet_names.append(name)

                    if len(name) > 31:
                        name = name[0:31]

                    replace_chars = ["/", "\\", "?", "*", ":", "[", "]"]
                    for char in replace_chars:
                        name = name.replace(char, "")
                    self.createSheetRawData(wb, raw_data["ColumnIDList"],
                                            name)

        newBytesIO = BytesIO()
        wb.save(newBytesIO)
        wb.close()
        blob = newBytesIO.getvalue()
        self.outputBytes = blob
        self.fs.put(blob, self.file_id)
        if self.save:
            Update.update("SpreadSheet", "LastUpdate", datetime.datetime.now(datetime.UTC),
                          self.spreadsheet["_id"],
                          get_db(), self.user_id)

    def remove_sheets(self, wb, change_string):
        sheets = wb.sheetnames
        if "Sheet" in sheets:
            sheet = wb["Sheet"]
            sheet.title = "Your workspace"

        if change_string in ["Metadata", "All"]:
            if "Experimental meta data" in sheets:
                sheet = wb["Experimental meta data"]
                max_row = sheet.max_row
                sheet.delete_rows(1, max_row)

            if "Sample meta data" in sheets:
                sheet = wb["Sample meta data"]
                max_row = sheet.max_row
                sheet.delete_rows(1, max_row)

            if "ResearchItem meta data" in sheets:
                sheet = wb["ResearchItem meta data"]
                max_row = sheet.max_row
                sheet.delete_rows(1, max_row)

            if "Combobox meta data" in sheets:
                sheet = wb["Combobox meta data"]
                max_row = sheet.max_row
                sheet.delete_rows(1, max_row)

        if change_string in ["RawData", "All"]:
            if "RawData" in sheets:
                sheet = wb["RawData"]
                wb.remove(sheet)

            for exp in self.exp_list:
                name = exp["Name"]
                if name in sheets:
                    sheet = wb[name]
                    wb.remove(sheet)

            for sheet_name in sheets:
                if sheet_name.startswith("Data table"):
                    sheet = wb[sheet_name]
                    wb.remove(sheet)

    def createSheetExperimentalMetaDataSheet(self, wb: openpyxl.Workbook, expList):
        if not expList:
            return
        sheets = wb.sheetnames
        if "Experimental meta data" in sheets:
            sheet = wb["Experimental meta data"]
        else:
            sheet = wb.create_sheet("Experimental meta data")
        sheet.protection.sheet = True
        sheet.protection.enable()

        row = 0
        for expDict in expList:
            sheet.append([expDict["Name"]])
            row += 1
            self.setRowBolded(sheet, row)

            for key in expDict:
                cellList = list()
                cellList.append(key)
                value = expDict[key]
                if isinstance(value, bson.ObjectId):
                    continue
                if isinstance(value, list):
                    for item in value:
                        cellList.append(item)
                else:
                    cellList.append(value)

                sheet.append(cellList)
                row += 1
            sheet.append([""])
            row += 1

        self.adjust_column_width(sheet)

        # sheet.column_dimensions["A"].width = 40
        # sheet.column_dimensions["B"].width = 80

    def createSheetSampleMetaDataSheet(self, wb: openpyxl.Workbook, sampleList):
        if not sampleList:
            return
        sheets = wb.sheetnames
        if "Sample meta data" in sheets:
            sheet = wb["Sample meta data"]
        else:
            sheet = wb.create_sheet("Sample meta data")
        sheet.protection.sheet = True
        sheet.protection.enable()

        row = 0
        for sampleList in sampleList:
            sheet.append([sampleList["Name"]])
            row += 1
            self.setRowBolded(sheet, row)

            for key in sampleList:
                cellList = list()
                cellList.append(key)
                value = sampleList[key]
                if isinstance(value, bson.ObjectId):
                    continue
                if isinstance(value, list):
                    for item in value:
                        cellList.append(item)
                else:
                    cellList.append(value)

                sheet.append(cellList)
                row += 1
            sheet.append([""])
            row += 1

        self.adjust_column_width(sheet)

        # sheet.column_dimensions["A"].width = 40
        # sheet.column_dimensions["B"].width = 80

    def createSheetResearchItemMetaDataSheet(self, wb: openpyxl.Workbook, riList):
        if not riList:
            return
        sheets = wb.sheetnames
        if "ResearchItem meta data" in sheets:
            sheet = wb["ResearchItem meta data"]
        else:
            sheet = wb.create_sheet("ResearchItem meta data")
        sheet.protection.sheet = True
        sheet.protection.enable()

        row = 0
        for ri in riList:
            sheet.append([ri["Name"]])
            row += 1
            self.setRowBolded(sheet, row)

            for key in ri:
                cellList = list()
                cellList.append(key)
                value = ri[key]
                if isinstance(value, bson.ObjectId):
                    continue
                if isinstance(value, list):
                    for item in value:
                        cellList.append(item)
                else:
                    cellList.append(value)

                sheet.append(cellList)
                row += 1
            sheet.append([""])
            row += 1

        self.adjust_column_width(sheet)
        # sheet.column_dimensions["A"].width = 40
        # sheet.column_dimensions["B"].width = 80

    def createSheetComboEntryMetaDataSheet(self, wb: openpyxl.Workbook, comboEntryDict):
        sheets = wb.sheetnames
        if "Combobox meta data" in sheets:
            sheet = wb["Combobox meta data"]
        else:
            sheet = wb.create_sheet("Combobox meta data")
        sheet.protection.sheet = True
        sheet.protection.enable()

        row = 1
        for key in comboEntryDict:
            if row > 1:
                sheet.append([])
                row += 1
            cellList = list()
            cellList.append(key)
            sheet.append(cellList)
            sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
            self.setRowBolded(sheet, row)
            row += 1
            fieldDataDict = comboEntryDict[key]
            for fieldKey in fieldDataDict:
                cellList = list()
                cellList.append(fieldKey)
                value = fieldDataDict[fieldKey]
                if isinstance(value, list):
                    for item in value:
                        cellList.append(item)
                else:
                    cellList.append(value)
                sheet.append(cellList)
                row += 1

        self.adjust_column_width(sheet)

    def createSheetRawData(self, wb: openpyxl.Workbook, columnIDList, name):
        sheets = wb.sheetnames
        if name in sheets:
            sheet = wb[name]
        else:
            sheet = wb.create_sheet(name)
        sheet.protection.sheet = True
        sheet.protection.enable()

        columnListInList = []
        for columnID in columnIDList:
            column = self.columnMapping[columnID]
            cellList = list()
            cellList.append(
                column["Name"]
            )
            data = column["Data"]
            if not data:
                columnListInList.append(cellList)
                continue
            if isinstance(data[0], list):
                # step based
                for step in data:
                    for item in step:
                        cellList.append(item)
                    cellList.append(None)
            else:
                for item in data:
                    cellList.append(item)
            columnListInList.append(cellList)

        rowsListInList = list(map(list, zip_longest(*columnListInList)))

        for row in rowsListInList:
            sheet.append(row)

        self.setRowBolded(sheet, 1)

    def setRowBolded(self, sheet: worksheet, rowNumber: int):
        rowTuple = sheet[rowNumber]
        for cell in rowTuple:
            font = Font(bold=True)
            cell.font = font

    def adjust_column_width(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        for column in sheet.columns:
            max_length = 0
            if isinstance(column[0], MergedCell):
                coordinate = column[0].coordinate
                pattern = r"[0-9]"
                column_letter = re.sub(pattern, "", coordinate)
                start_row = 0
            else:
                column_letter = column[0].column_letter
                start_row = 0
            for cell in column[start_row:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def createListExp(self):
        from tenjin.mongo_engine.Link import Link
        returnListExp = []
        for exp in self.exp_list:
            """Exp: Name, FieldData, UsedSamples, CreatedSamples, Groups"""
            returnDict = dict()
            returnDict["Attached samples"] = []
            returnDict["Linked items"] = []
            returnDict["Groups"] = []

            returnDict["Name"] = exp["Name"]

            """groups"""
            groupIDList = exp["GroupIDList"]
            findDict = {
                "_id": {
                    "$in": groupIDList
                }
            }
            groupList = self.getData("Group", findDict)
            for group in groupList:
                returnDict["Groups"].append(group["Name"])

            """used samples """
            links = Link.objects(Q(DataID1=exp["_id"]) | Q(DataID2=exp["_id"]))
            sampleIDList = []
            itemIDList = []
            for link in links:
                if link.DataID1 == exp["_id"]:
                    if link.Collection2 == "Sample":
                        sampleIDList.append(link.DataID2)
                    else:
                        itemIDList.append(link.DataID2)
                elif link.DataID2 == exp["_id"]:
                    if link.Collection1 == "Sample":
                        sampleIDList.append(link.DataID1)
                    else:
                        itemIDList.append(link.DataID1)

            sampleList = self.getData("Sample", findDict)
            for sample in sampleList:
                returnDict["Attached samples"].append(sample["Name"])
            itemList = self.getData("ResearchItem", findDict)
            for item in itemList:
                returnDict["Linked items"].append(item["Name"])

            # Create fieldData
            fieldDict = self.createDictFieldData(exp)
            returnDict.update(fieldDict)

            returnListExp.append(returnDict)

        return returnListExp

    def createListSample(self):
        returnListSample = []
        for sample in self.sample_list:
            returnDict = dict()
            returnDict["Groups"] = []
            returnDict["Attached to exp"] = []

            returnDict["Name"] = sample["Name"]
            links = Link.objects(
                (Q(DataID1=sample["_id"]) & Q(Collection2="Experiment")) |
                (Q(DataID2=sample["_id"]) & Q(Collection1="Experiment"))
            )
            exp_id_list = []
            for link in links:
                if link.DataID1 == sample["_id"]:
                    exp_id_list.append(link.DataID2)
                else:
                    exp_id_list.append(link.DataID1)

            """Used by exp """
            findDict = {
                "_id": {
                    "$in": [exp_id_list]
                }
            }
            expList = self.getData("Experiment", findDict)
            for exp in expList:
                returnDict["Attached to exp"].append(exp["Name"])

            """groups"""
            groupIDList = sample["GroupIDList"]
            findDict = {
                "_id": {
                    "$in": groupIDList
                }
            }
            groupList = self.getData("Group", findDict)
            for group in groupList:
                returnDict["Groups"].append(group["Name"])

            # Create fieldData
            fieldDict = self.createDictFieldData(sample)
            returnDict.update(fieldDict)

            returnListSample.append(returnDict)

        return returnListSample

    def createListResearchItems(self):
        returnListResearchItems = []

        for ri in self.research_item_list:
            returnDict = dict()
            returnDict["Groups"] = []
            returnDict["Attached to exp"] = []
            returnDict["Category"] = None

            returnDict["Name"] = ri["Name"]

            """Used by exp """
            findDict = {
                "ResearchItemIDList": {
                    "$in": [ri["_id"]]
                }
            }
            expList = self.getData("Experiment", findDict)
            for exp in expList:
                returnDict["Attached to exp"].append(exp["Name"])

            """groups"""
            groupIDList = ri["GroupIDList"]
            findDict = {
                "_id": {
                    "$in": groupIDList
                }
            }
            groupList = self.getData("Group", findDict)
            for group in groupList:
                returnDict["Groups"].append(group["Name"])

            """Category"""
            category_id = ri["ResearchCategoryID"]
            findDict = {
                "_id": category_id
            }
            category: dict = self.getData("ResearchCategory", findDict, findOne=True)
            returnDict["Category"] = category["Name"]

            # Create fieldData
            fieldDict = self.createDictFieldData(ri)
            returnDict.update(fieldDict)

            returnListResearchItems.append(returnDict)

        return returnListResearchItems

    def createDictComboData(self):
        returnDict = dict()
        for comboEntryID in self.comboBoxEntryMapping:
            comboEntry = self.comboBoxEntryMapping[comboEntryID]
            fieldDict = self.createDictFieldData(
                comboEntry)
            returnDict[comboEntry["Name"]] = fieldDict

        return returnDict

    def createDictFieldData(self, dataDict):

        returnDict = dict()

        fieldDataIDList = dataDict["FieldDataIDList"]

        # build the return dict by iterating the fieldDataList. The key of the return dict is the field name and the value
        # is the value of the field data. for combobox entries it is the name, for notebooks the content, for links the name
        # of the target and for rawDataCalc the result. For numeric fields the fieldName is extended by the unit short name
        for fieldDataID in fieldDataIDList:
            fieldData = self.fieldDataMapping[fieldDataID]
            field = self.fieldMapping[fieldData["FieldID"]]
            if fieldData["Value"] is None:
                returnDict[field["Name"]] = ""
                continue
            fieldType = field["Type"]
            fieldName = field["Name"]
            unitID = fieldData["UnitID"]
            value = fieldData["Value"]
            if fieldType in ["ComboBoxSynonym", "ComboBox"]:
                if value is not None:
                    value = self.comboBoxEntryMapping[value]["Name"]
            elif fieldType == "Link":
                if value is not None:
                    link = self.linkMapping[value]
                    linkTargetID = link["TargetID"]
                    target = self.linkTargetDataMapping[linkTargetID]
                    value = target["Name"]
            elif fieldType == "MultiLine":
                if value is not None:
                    value = ""
                    notebook: dict = self.notebookMapping.get(value)
                    if notebook:
                        value = notebook.get("Content", "")
            elif fieldType == "RawDataCalc":
                if value is not None:
                    calculation = self.calculationMapping[value]
                    result = calculation["Output"]
                    value = []
                    for key in result:
                        value.append(key)
                        value.append(result[key])
                pass
            elif fieldType == "Numeric":
                if unitID is not None:
                    unit = self.unitMapping[unitID]
                    value = [value, unit["ShortName"]]
            returnDict[fieldName] = value

        return returnDict

    def getRawData(self):

        findDict = {"$or": [
            {"ExpID": {"$in": self.exp_id_list}},
            {"SampleID": {"$in": self.sample_id_list}},
            {"ResearchItemID": {"$in": self.research_item_id_list}}
        ]
        }
        rawDataList = list(self.getData("RawData", findDict))
        if rawDataList:
            self.rawDataList = rawDataList
            self.rawDataIDList = [rawData["_id"] for rawData in rawDataList]
            for raw_data in rawDataList:
                if raw_data["ExpID"]:
                    if raw_data["ExpID"] not in self.exp_id_rawdata_mapping:
                        self.exp_id_rawdata_mapping[raw_data["ExpID"]] = []
                    self.exp_id_rawdata_mapping[raw_data["ExpID"]].append(raw_data)
                elif raw_data["SampleID"]:
                    if raw_data["SampleID"] not in self.sample_id_rawdata_mapping:
                        self.sample_id_rawdata_mapping[raw_data["SampleID"]] = []
                    self.sample_id_rawdata_mapping[raw_data["SampleID"]].append(raw_data)
                elif raw_data["ResearchItemID"]:
                    if raw_data["ResearchItemID"] not in self.research_item_id_rawdata_mapping:
                        self.research_item_id_rawdata_mapping[raw_data["ResearchItemID"]] = []
                    self.research_item_id_rawdata_mapping[raw_data["ResearchItemID"]].append(raw_data)

            for rawData in self.rawDataList:
                self.columnIDList.extend(rawData["ColumnIDList"])

    def getRawDataColumnList(self):

        findDict = {
            "_id": {"$in": self.columnIDList}
        }
        columnList = self.getData("Column", findDict)
        self.columnList = columnList
        self.columnMapping = {column["_id"]: column for column in columnList}

    def getAllFieldDataList(self):
        # get all field data
        fieldDataIDList = []
        for exp in self.exp_list:
            fieldDataIDList.extend(exp["FieldDataIDList"])
        for sample in self.sample_list:
            fieldDataIDList.extend(sample["FieldDataIDList"])
        for researchitem in self.research_item_list:
            fieldDataIDList.extend(researchitem["FieldDataIDList"])
        if not fieldDataIDList:
            return
        findDict = {
            "_id": {
                "$in": fieldDataIDList
            }
        }
        self.fieldDataIDList = fieldDataIDList
        self.fieldDataList = self.getData("FieldData", findDict)
        self.fieldDataMapping = {fieldData["_id"]: fieldData for fieldData in self.fieldDataList}

    def getExpFieldList(self):
        # get all needed fields from the fieldDataList to obtain the field types and names
        fieldIDList = [fieldData["FieldID"]
                       for fieldData in self.fieldDataList]
        if not fieldIDList:
            return
        findDict = {
            "_id": {
                "$in": fieldIDList
            }
        }
        self.fieldList = self.getData("Field", findDict)
        self.fieldMapping = {field["_id"]: field for field in self.fieldList}

    def getComboBoxEntryFieldDataAndFieldRecursive(self, fieldDataIDList):
        missingFieldDataID = []
        for fieldDataID in fieldDataIDList:
            if fieldDataID not in self.fieldDataMapping.keys():
                missingFieldDataID.append(fieldDataID)
        if missingFieldDataID:
            findDict = {"_id": {"$in": missingFieldDataID}}
            missingFieldDataList = self.getData("FieldData", findDict)
            for fieldData in missingFieldDataList:
                self.fieldDataMapping[fieldData["_id"]] = fieldData
        missingFieldIDs = []
        missingUnitIDs = []
        for fieldDataID in fieldDataIDList:
            fieldData = self.fieldDataMapping[fieldDataID]
            fieldID = fieldData["FieldID"]
            unitID = fieldData["UnitID"]
            if not fieldID in self.fieldMapping.keys():
                missingFieldIDs.append(fieldID)
            if not unitID in self.unitMapping.keys():
                missingUnitIDs.append(unitID)
        if missingFieldIDs:
            findDict = {"_id": {"$in": missingFieldIDs}}
            fieldList = self.getData("Field", findDict)
            for field in fieldList:
                self.fieldMapping[field["_id"]] = field

        comboBoxEntryIDList = []
        for fieldDataID in fieldDataIDList:
            fieldData = self.fieldDataMapping[fieldDataID]
            fieldID = fieldData["FieldID"]
            field = self.fieldMapping[fieldID]
            fieldType = field["Type"]
            if fieldType in ["ComboBoxSynonym", "ComboBox"]:
                if fieldData["Value"] is not None:
                    comboBoxEntryIDList.append(fieldData["Value"])

        missingComboBoxEntries = []
        for comboBoxEntryID in comboBoxEntryIDList:
            if comboBoxEntryID not in self.comboBoxEntryMapping.keys():
                missingComboBoxEntries.append(comboBoxEntryID)

        # Obtain all used comboBoxEntries
        findDict = {"_id": {"$in": missingComboBoxEntries}}
        missingComboBoxEntryList = self.getData("ComboBoxEntry", findDict)
        newFieldDataIDList = []
        for comboBoxEntry in missingComboBoxEntryList:
            self.comboBoxEntryMapping[comboBoxEntry["_id"]] = comboBoxEntry
            comboFieldDataIDList = comboBoxEntry["FieldDataIDList"]
            for fieldDataID in comboFieldDataIDList:
                newFieldDataIDList.append(fieldDataID)
        if newFieldDataIDList:
            self.getComboBoxEntryFieldDataAndFieldRecursive(newFieldDataIDList)

    def getAllUnit(self):
        # get all units for all numeric fields
        unitIDList = []
        for fieldData in self.fieldDataMapping.values():
            unitID = fieldData["UnitID"]
            if not unitID:
                continue
            if unitID not in unitIDList:
                unitIDList.append(unitID)

        if not unitIDList:
            return
        findDict = {"_id": {"$in": unitIDList}}
        unitList = self.getData("Unit", findDict)
        self.unitMapping = {unit["_id"]: unit for unit in unitList}

    def getAllNotebook(self):
        # get all notebooks for all numeric fields
        notebookIDList = []
        for fieldData in self.fieldDataMapping.values():
            field = self.fieldMapping[fieldData["FieldID"]]
            if field["Type"] != "MultiLine":
                continue
            notebookID = fieldData["Value"]
            if not notebookID:
                continue
            if notebookID not in notebookIDList:
                notebookIDList.append(notebookID)

        if not notebookIDList:
            return
        findDict = {"_id": {"$in": notebookIDList}}
        notebookList = self.getData("Notebook", findDict)
        self.notebookMapping = {notebook["_id"]: notebook for notebook in notebookList}

    def getAllCalculation(self):
        calculationIDList = []
        for fieldData in self.fieldDataMapping.values():
            field = self.fieldMapping[fieldData["FieldID"]]
            if field["Type"] != "RawDataCalc":
                continue
            calculationID = fieldData["Value"]
            if not calculationID:
                continue
            if not calculationID in calculationIDList:
                calculationIDList.append(calculationID)

        if not calculationIDList:
            return
        findDict = {"_id": {"$in": calculationIDList}}
        calculationList = self.getData("Calculation", findDict)
        self.calculationMapping = {calculation["_id"]: calculation for calculation in calculationList}

    def getAllLink(self):
        # get all notebooks for all numeric fields
        linkIDList = []
        for fieldData in self.fieldDataMapping.values():
            field = self.fieldMapping[fieldData["FieldID"]]
            if field["Type"] != "MultiLine":
                continue
            linkID = fieldData["Value"]
            if not linkID:
                continue
            if linkID not in linkIDList:
                linkIDList.append(linkID)

        if not linkIDList:
            return
        findDict = {"_id": {"$in": linkIDList}}
        linkList = self.getData("FieldDataLink", findDict)
        self.linkMapping = {link["_id"]: link for link in linkList}

        linkCollectionDict = {}
        for link in linkList:
            targetCollection = link["TargetCollection"]
            targetID = link["TargetID"]
            if targetCollection not in linkCollectionDict:
                linkCollectionDict[targetCollection] = []
            linkCollectionDict[targetCollection].append(targetID)
        linkTargetDataList = []
        for targetCollection in linkCollectionDict:
            findDict = {
                "_id": {
                    "$in": linkCollectionDict[targetCollection]
                }
            }
            data = self.getData(targetCollection, findDict)
            if data:
                linkTargetDataList.extend(data)
        self.linkTargetDataMapping = {
            linkTargetData["_id"]: linkTargetData for linkTargetData in linkTargetDataList}

    def get_dataview_data(self):
        from tenjin.web.helper.fieldsHelper import get_field_data_values
    
        item_id_list = self.dataview["ItemIDList"]
        extra_items = []
        total_linked_mapping = {}
        async def _get_linked_items_recursive():
            coroutines = [get_linked_items_recursive(
                self.dataview["IncludeLinked"], item_id_list, retrieve_linked_mapping=True)]
            result = await asyncio.gather(*coroutines)
            extra_items.extend(result[0][0])
            total_linked_mapping.update(result[0][1])
        
        
        asyncio.run(_get_linked_items_recursive())
        number_of_layer = 0
        for item_id in total_linked_mapping:
            number_of_layer = len(total_linked_mapping[item_id])
            break
        
        extra_items.extend(item_id_list)
        items = get_item_for_dashboard(self.dataview["ProjectID"], extra_items)
        item_mapping = {item["id"]: item for item in items}

        displayed_columns = self.dataview["DisplayedColumns"]
        find_dict = {"Name": {"$in": displayed_columns},
                     "ProjectID": self.dataview["ProjectID"]}
        fields = []
        for std_field in self.std_fields_dataview:
            if std_field in displayed_columns:
                fields.append({"Name": std_field})

        fields.extend(self.getData("Field", find_dict))

        raw_data_calc_length_mapping = {}
        raw_data_calc_column_name_mapping = {}
        item_rows_length_mapping = {}

        field_id_list = []
        for field in fields:
            field_id = field.get("_id", None)
            if field_id is None:
                continue
            field_id_list.append(field_id)

        field_data_id_list = []
        for item in items:
            fd_list = item["field_data_id_list"]
            fd_list = [bson.ObjectId(f) for f in fd_list]
            field_data_id_list.extend(fd_list)

        field_data_mapping = {}
        s = time.time()
        for i in range(0, len(field_data_id_list), 20000):
            start = i
            end = i + 20000
            if end > len(field_data_id_list):
                end = len(field_data_id_list)
            _field_data_id_list = field_data_id_list[start:end]
            _field_data_mapping = get_field_data_values(_field_data_id_list, field_id_list)
            field_data_mapping.update(_field_data_mapping)
        print("get field_data", time.time() - s)
        
        for item_id in item_id_list:
            item = item_mapping.get(str(item_id), None)
            for layer_number in range(0, number_of_layer + 1): 
                if layer_number == 0:
                    linked_items = [item_id]
                else:
                    linked_items = total_linked_mapping[item_id][layer_number - 1] 
        
                for field in fields:
                    field_id = field.get("_id", None)
                    if not item.get(field_id):
                        if item["id"] not in item_rows_length_mapping:
                            item_rows_length_mapping[item["id"]] = 1
                        if field_id is None:
                            if field.get("Name") == "Name":
                                item["Name"] = [{"value": item["name"]}]
                            elif field.get("Name") == "Short ID":
                                item["Short ID"] = [{"value": item["short_id"]}]
                            elif field.get("Name") == "Type":
                                item["Type"] = [{"value": item["type"]}]
                        else:
                            for _id in linked_items:
                                _item = item_mapping.get(str(_id), None)
                                field_data_id_list = _item["field_data_id_list"]
                                field_data_id_list = [bson.ObjectId(f) for f in field_data_id_list]
                                # field_data_mapping = get_field_data_values(field_data_id_list, [field_id])

                                for field_data_id in field_data_id_list:
                                    fd = field_data_mapping.get(field_data_id, None)
                                    if fd:
                                        if fd.get("field_id") == str(field_id):
                                            if field_id not in item:
                                                item[field_id] = []

                                            item[field_id].append(fd)
                    
                        if field_id in item:
                            current_length = item_rows_length_mapping.get(item["id"])
                            if len(item[field_id]) > current_length:
                                item_rows_length_mapping[item["id"]] = len(item[field_id])

                            if field.get("Type") == "RawDataCalc":
                                if not item[field_id]:
                                    if field_id not in raw_data_calc_length_mapping:
                                        raw_data_calc_length_mapping[field_id] = 1
                                        raw_data_calc_column_name_mapping[field_id] = [None]
                                    continue
                                field_data = item[field_id][0]
                                value = field_data["original_value"]
                                if value:
                                    length = len(value.keys())
                                    if field_id in raw_data_calc_length_mapping:
                                        if length > raw_data_calc_length_mapping[field_id]:
                                            raw_data_calc_length_mapping[field_id] = length
                                            raw_data_calc_column_name_mapping[field_id] = list(value.keys())
                                    else:
                                        raw_data_calc_length_mapping[field_id] = len(value.keys())
                                        raw_data_calc_column_name_mapping[field_id] = list(value.keys())

        
        
        # for field in fields:
        #     field_id = field.get("_id", None)
        #     for item_id in item_id_list:
        #         item = item_mapping.get(str(item_id), None)
        #         if "/sep/" in item.get("id", ""):
        #             continue
        #         if item["id"] not in item_rows_length_mapping:
        #             item_rows_length_mapping[item["id"]] = 1
        #         if field_id is None:
        #             if field.get("Name") == "Name":
        #                 item["Name"] = [{"value": item["name"]}]
        #             elif field.get("Name") == "Short ID":
        #                 item["Short ID"] = [{"value": item["short_id"]}]
        #             elif field.get("Name") == "Type":
        #                 item["Type"] = [{"value": item["type"]}]
        #         else:
        #             field_data_id_list = item["field_data_id_list"]
        #             field_data_id_list = [bson.ObjectId(f) for f in field_data_id_list]
        #             # field_data_mapping = get_field_data_values(field_data_id_list, [field_id])

        #             for field_data_id in field_data_id_list:
        #                 fd = field_data_mapping.get(field_data_id, None)
        #                 if fd:
        #                     if fd.get("field_id") == str(field_id):
        #                         if field_id not in item:
        #                             item[field_id] = []

        #                         item[field_id].append(fd)

        #             if field_id in item:
        #                 current_length = item_rows_length_mapping.get(item["id"])
        #                 if len(item[field_id]) > current_length:
        #                     item_rows_length_mapping[item["id"]] = len(item[field_id])

        #                 if field.get("Type") == "RawDataCalc":
        #                     if not item[field_id]:
        #                         if field_id not in raw_data_calc_length_mapping:
        #                             raw_data_calc_length_mapping[field_id] = 1
        #                             raw_data_calc_column_name_mapping[field_id] = [None]
        #                         continue
        #                     field_data = item[field_id][0]
        #                     value = field_data["original_value"]
        #                     if value:
        #                         length = len(value.keys())
        #                         if field_id in raw_data_calc_length_mapping:
        #                             if length > raw_data_calc_length_mapping[field_id]:
        #                                 raw_data_calc_length_mapping[field_id] = length
        #                                 raw_data_calc_column_name_mapping[field_id] = list(value.keys())
        #                         else:
        #                             raw_data_calc_length_mapping[field_id] = len(value.keys())
        #                             raw_data_calc_column_name_mapping[field_id] = list(value.keys())

        column = 1
        for field in fields:
            if not field.get("Type"):
                field["start_column"] = column
                field["end_column"] = column + 1 - 1
                column += 1
                continue
            if field.get("Type") == "Numeric":
                field["start_column"] = column
                field["end_column"] = column + 3 - 1
                column += 3
            elif field.get("Type") == "RawDataCalc":
                length = raw_data_calc_length_mapping.get(field["_id"], 1)
                field["start_column"] = column
                field["end_column"] = column + length - 1
                column += length
                field["sub_column_names"] = raw_data_calc_column_name_mapping[field["_id"]]
            else:
                field["start_column"] = column
                field["end_column"] = column + 1 - 1
                column += 1
                continue

        row = 2
        items = [item_mapping.get(str(item_id), None) for item_id in item_id_list]
        for item in items:
            if "/sep/" in item.get("id", ""):
                continue
            length = item_rows_length_mapping.get(item["id"])
            item["start_row"] = row
            item["end_row"] = row + length - 1
            row += length

        return items, fields

    def create_workbook_for_dataview(self):
        wb = openpyxl.Workbook()
        sheet = wb["Sheet"]
        sheet.title = "Your workspace"

        self.create_dataview_sheet(wb)
        output_io = BytesIO()
        wb.save(output_io)
        self.outputBytes = output_io.getvalue()

        file_name = f"{self.dataview.get('Name', 'DataView')}_spreadsheet.xlsx"
        self.file_id = self.fs.put(self.outputBytes, fileName=file_name)
        spreadsheetDict = {
            "LastUpdate": datetime.datetime.now(datetime.UTC),
            "FileID": self.file_id,
            "DataViewID": self.dataview_id,
            "MD5": self.md5
        }
        _id = Create.create("SpreadSheet", spreadsheetDict, get_db(), self.user_id)
        self.spreadsheet_id = _id
        self.spreadsheet = get_db().db["SpreadSheet"].find_one({"_id": _id})

    def update_workbook_for_dataview(self):
        spreadsheet_file_id = self.spreadsheet["FileID"]
        spreadsheet_bytes = self.fs.get_file(spreadsheet_file_id)
        spreadsheet_bytes_io = BytesIO(spreadsheet_bytes)
        wb = openpyxl.load_workbook(spreadsheet_bytes_io)
        sheets = wb.sheetnames
        for sheetName in sheets:
            if sheetName == "Your workspace":
                continue
            try:
                sheet = wb[sheetName]
                wb.remove(sheet)
            except:
                pass
        self.create_dataview_sheet(wb)
        output_io = BytesIO()
        wb.save(output_io)
        self.outputBytes = output_io.getvalue()
        self.fs.put(self.outputBytes, self.file_id)
        Update.update("SpreadSheet", "LastUpdate",
                      datetime.datetime.now(datetime.UTC), self.spreadsheet["_id"],
                      get_db(), self.user_id)

    def create_dataview_sheet(self, wb: openpyxl.Workbook):
        name = self.dataview.get("Name", "Data")
        sheet = wb.create_sheet(name)
        sheet.protection.sheet = True
        sheet.protection.enable()

        items, fields = self.get_dataview_data()
        row = 1
        additional_header = False
        for field in fields:
            start_column = field["start_column"]
            end_column = field["end_column"]
            if end_column != start_column:
                sheet.merge_cells(start_row=row, end_row=row,
                                  start_column=start_column, end_column=end_column)
            sheet.cell(row=row, column=start_column).value = field["Name"]
            if field.get("Type") == "Numeric":
                additional_header = True
                sheet.cell(row=row + 1, column=start_column).value = "Value"
                sheet.cell(row=row + 1, column=start_column + 1).value = "Unit"
                sheet.cell(row=row + 1, column=start_column + 2).value = "SI-Value"
            elif field.get("Type") == "RawDataCalc":
                additional_header = True
                sub_column_names = field["sub_column_names"]
                for pos, sub_column_name in enumerate(sub_column_names):
                    sheet.cell(row=row + 1, column=start_column + pos).value = sub_column_name
            self.setRowBolded(sheet, row)

        numeric_keys = ["original_value", "unit", "si_value"]
        for item in items:
            if "/sep/" in item.get("id", ""):
                continue
            start_row = item["start_row"]
            end_row = item["end_row"]
            if additional_header:
                start_row += 1
                end_row += 1
            number_rows = end_row - start_row + 1

            for field in fields:
                start_column = field["start_column"]
                end_column = field["end_column"]
                number_columns = end_column - start_column + 1
                field_name = field["Name"]
                field_type = field.get("Type")

                if field_name in self.std_fields_dataview:
                    field_data_list = item[field_name]
                else:
                    field_id = field.get("_id")
                    field_data_list = item.get(field_id)

                if not field_data_list:
                    continue
                if field_type == "Numeric":
                    if len(field_data_list) == 1:
                        field_data = field_data_list[0]
                        if number_rows > 1:
                            for pos, key in enumerate(numeric_keys):
                                c = start_column + pos
                                sheet.merge_cells(start_row=start_row, end_row=end_row,
                                                  start_column=c, end_column=c)
                                sheet.cell(row=start_row, column=c).value = field_data[key]
                        else:
                            for pos, key in enumerate(numeric_keys):
                                c = start_column + pos
                                sheet.cell(row=start_row, column=c).value = field_data[key]
                    else:
                        for pos_row, fd in enumerate(field_data_list):
                            row = start_row + pos_row
                            for pos_col, key in enumerate(numeric_keys):
                                c = start_column + pos_col
                                sheet.cell(row=row, column=c).value = fd[key]

                elif field_type == "RawDataCalc":
                    sub_column_names = field.get("sub_column_names")
                    if sub_column_names:
                        if len(field_data_list) == 1:
                            field_data = field_data_list[0]
                            if number_rows > 1:
                                for pos, sub_column_name in enumerate(sub_column_names):
                                    c = start_column + pos
                                    sheet.merge_cells(start_row=start_row, end_row=end_row,
                                                      start_column=c, end_column=c)
                                    if isinstance(field_data["original_value"], dict):
                                        sheet.cell(row=start_row, column=c).value = field_data["original_value"].get(
                                            sub_column_name)
                            else:
                                for pos, sub_column_name in enumerate(sub_column_names):
                                    c = start_column + pos
                                    if isinstance(field_data["original_value"], dict):
                                        sheet.cell(row=start_row, column=c).value = field_data["original_value"].get(
                                            sub_column_name)
                        else:
                            for pos_row, fd in enumerate(field_data_list):
                                row = start_row + pos_row
                                for pos_col, sub_column_name in enumerate(sub_column_names):
                                    c = start_column + pos_col
                                    if isinstance(fd["original_value"], dict):
                                        sheet.cell(row=row, column=c).value = fd["original_value"].get(
                                            sub_column_name)
                else:
                    if len(field_data_list) == 1:
                        if number_rows > 1:
                            sheet.merge_cells(start_row=start_row, end_row=end_row,
                                              start_column=start_column, end_column=end_column)
                        field_data = field_data_list[0]
                        sheet.cell(row=start_row, column=start_column).value = field_data["value"]
                    else:
                        for pos, fd in enumerate(field_data_list):
                            row = start_row + pos
                            sheet.cell(row=row, column=start_column).value = fd["value"]

        self.adjust_column_width(sheet)


if __name__ == "__main__":
    from tenjin import create_app

    app = create_app(minimal=True)
    import flask

    with app.app_context():
        dataviewid = ObjectId("64cb425559df439d2ffe2064")
        dataview = get_db().db["DataView"].find_one({"_id": dataviewid})
        user_id = ObjectId("5b1811ffdc92661b604be635")
        flask.g.user = user_id
        spreadsheet_gen = SpreadsheetGenerator(dataview=dataview, user_id=user_id)
        with open("spreadsheet.xlsx", "wb+") as f:
            f.write(spreadsheet_gen.outputBytes)
